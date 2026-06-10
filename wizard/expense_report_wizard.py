import io
import base64
from datetime import date

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


class ExpenseReportWizard(models.TransientModel):
    _name = "business.expense.report.wizard"
    _description = "Expense Report Wizard"

    date_from = fields.Date(
        string="Start Date",
        required=True,
        default=lambda self: date.today().replace(day=1),
    )
    date_to = fields.Date(
        string="End Date",
        required=True,
        default=lambda self: date.today(),
    )
    report_type = fields.Selection(
        selection=[
            ("summary", "Summary Only"),
            ("detailed", "Detailed Report"),
        ],
        string="Report Type",
        default="detailed",
        required=True,
    )
    expense_ids = fields.Many2many(
        comodel_name="business.expense",
        string="Expenses",
        compute="_compute_expenses",
        store=False,
    )
    category_id = fields.Many2one(
        comodel_name="business.expense.category",
        string="Category",
    )
    account_id = fields.Many2one(
        comodel_name="business.expense.account",
        string="Payment Account",
        check_company=True,
    )
    total_with_tax = fields.Monetary(
        string="Total Paid",
        compute="_compute_totals",
        currency_field="currency_id",
    )
    total_tax = fields.Monetary(
        string="Total Tax",
        compute="_compute_totals",
        currency_field="currency_id",
    )
    total_without_tax = fields.Monetary(
        string="Total Subtotal",
        compute="_compute_totals",
        currency_field="currency_id",
    )
    expense_count = fields.Integer(
        string="Number of Expenses",
        compute="_compute_totals",
    )
    currency_id = fields.Many2one(
        comodel_name="res.currency",
        string="Currency",
        related="company_id.currency_id",
        readonly=True,
    )
    company_id = fields.Many2one(
        comodel_name="res.company",
        string="Company",
        required=True,
        default=lambda self: self.env.company,
    )

    @api.depends("date_from", "date_to", "category_id", "account_id", "company_id")
    def _compute_expenses(self):
        for wizard in self:
            if wizard.date_from and wizard.date_to:
                domain = [
                    ("date", ">=", wizard.date_from),
                    ("date", "<=", wizard.date_to),
                    ("company_id", "=", wizard.company_id.id),
                ]
                if wizard.category_id:
                    domain.append(("category_id", "child_of", wizard.category_id.id))
                if wizard.account_id:
                    domain.append(("account_id", "=", wizard.account_id.id))
                wizard.expense_ids = self.env["business.expense"].search(
                    domain,
                    order="date desc",
                )
            else:
                wizard.expense_ids = False

    @api.depends("expense_ids")
    def _compute_totals(self):
        for wizard in self:
            expenses = wizard.expense_ids
            wizard.total_with_tax = sum(expenses.mapped("total_with_tax"))
            wizard.total_tax = sum(expenses.mapped("tax_amount"))
            wizard.total_without_tax = sum(expenses.mapped("total_without_tax"))
            wizard.expense_count = len(expenses)

    @api.onchange("date_from")
    def _onchange_date_from(self):
        if self.date_from and self.date_to and self.date_from > self.date_to:
            self.date_to = self.date_from

    @api.constrains("date_from", "date_to")
    def _check_dates(self):
        for wizard in self:
            if wizard.date_from > wizard.date_to:
                raise UserError(_("Start date must be before or equal to end date."))

    def action_generate_pdf(self):
        self.ensure_one()
        return self.env.ref(
            "business_expense.action_report_business_expense"
        ).report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(
                _(
                    "The openpyxl library is not installed. Please contact your system administrator."
                )
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Expense Report"

        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font_white = Font(bold=True, color="FFFFFF")

        company = self.company_id

        sheet["A1"] = "Business Expense Report"
        sheet["A1"].font = title_font
        sheet.merge_cells("A1:H1")

        sheet["A2"] = f"Period: {self.date_from} to {self.date_to}"
        sheet.merge_cells("A2:H2")

        sheet["A3"] = f"Company: {company.name}"
        sheet.merge_cells("A3:H3")

        sheet["A5"] = "Summary"
        sheet["A5"].font = header_font

        sheet["A6"] = "Total Expenses:"
        sheet["B6"] = self.expense_count
        sheet["A7"] = "Total Paid:"
        sheet["B7"] = self.total_with_tax
        sheet["B7"].number_format = "#,##0.00"
        sheet["A8"] = "Total Tax:"
        sheet["B8"] = self.total_tax
        sheet["B8"].number_format = "#,##0.00"
        sheet["A9"] = "Total Subtotal:"
        sheet["B9"] = self.total_without_tax
        sheet["B9"].number_format = "#,##0.00"

        if self.report_type == "detailed":
            start_row = 11
            headers = [
                "Date",
                "Expense Name",
                "Category",
                "Payment Account",
                "Subtotal",
                "Total Paid",
                "Tax Paid",
                "Created By",
            ]
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=start_row, column=col)
                cell.value = header
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            row = start_row + 1
            for expense in self.expense_ids:
                sheet.cell(row=row, column=1).value = str(expense.date)
                sheet.cell(row=row, column=2).value = expense.name
                sheet.cell(row=row, column=3).value = expense.category_id.display_name
                sheet.cell(row=row, column=4).value = expense.account_id.display_name
                sheet.cell(row=row, column=5).value = expense.total_without_tax
                sheet.cell(row=row, column=5).number_format = "#,##0.00"
                sheet.cell(row=row, column=6).value = expense.total_with_tax
                sheet.cell(row=row, column=6).number_format = "#,##0.00"
                sheet.cell(row=row, column=7).value = expense.tax_amount or 0
                sheet.cell(row=row, column=7).number_format = "#,##0.00"
                sheet.cell(row=row, column=8).value = expense.user_id.name or ""

                for col in range(1, 9):
                    sheet.cell(row=row, column=col).border = border
                row += 1

        column_widths = [
            (1, 12),
            (2, 40),
            (3, 24),
            (4, 28),
            (5, 15),
            (6, 15),
            (7, 15),
            (8, 20),
        ]
        for col, width in column_widths:
            sheet.column_dimensions[get_column_letter(col)].width = width

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = f"expense_report_{self.date_from}_{self.date_to}.xlsx"

        attachment = self.env["ir.attachment"].create(
            {
                "name": filename,
                "type": "binary",
                "datas": base64.b64encode(buffer.read()),
                "res_model": self._name,
                "res_id": self.id,
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )

        return {
            "type": "ir.actions.act_window",
            "name": _("Excel Export"),
            "res_model": "ir.attachment",
            "res_id": attachment.id,
            "view_mode": "form",
            "views": [(False, "form")],
            "target": "new",
        }
